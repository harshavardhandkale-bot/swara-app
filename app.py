import streamlit as st
import pandas as pd
import json
import os
import base64
from datetime import datetime, date
from io import BytesIO
from PIL import Image
import google.generativeai as genai
from pydantic import BaseModel, Field
from typing import Optional
import sqlite3

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Swara — Kale Pharma Pvt Ltd",
    page_icon="💊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .company-header {
        font-size: 0.75rem;
        font-weight: 700;
        letter-spacing: 0.15em;
        text-transform: uppercase;
        color: #17533a;
        margin-bottom: 0;
    }
    .page-title {
        font-size: 2rem;
        font-weight: 700;
        color: #1a1a1a;
        margin-top: 0;
    }
    .metric-card {
        background: #f5f7f5;
        border-radius: 10px;
        padding: 1rem 1.5rem;
        border-left: 4px solid #17533a;
    }
    .stButton > button {
        background-color: #17533a;
        color: white;
        border: none;
        border-radius: 8px;
        padding: 0.5rem 1.5rem;
        font-weight: 600;
    }
    .stButton > button:hover {
        background-color: #0f3d2a;
        color: white;
    }
</style>
""", unsafe_allow_html=True)

# ── Gemini API setup ──────────────────────────────────────────────────────────
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)

# ── Database setup ────────────────────────────────────────────────────────────
DB_PATH = "swara_expenses.db"

def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            expense_date TEXT NOT NULL,
            particulars TEXT NOT NULL,
            vendor TEXT,
            category TEXT NOT NULL,
            amount REAL NOT NULL,
            payment_mode TEXT DEFAULT 'Cash',
            entered_by TEXT DEFAULT 'Owner',
            notes TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.commit()
    conn.close()

def get_all_expenses():
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query(
        "SELECT * FROM expenses ORDER BY expense_date DESC, created_at DESC",
        conn
    )
    conn.close()
    return df

def add_expense(expense_date, particulars, vendor, category, amount, payment_mode, entered_by, notes):
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        INSERT INTO expenses (expense_date, particulars, vendor, category, amount, payment_mode, entered_by, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (expense_date, particulars, vendor, category, amount, payment_mode, entered_by, notes))
    conn.commit()
    conn.close()

def delete_expense(expense_id):
    conn = sqlite3.connect(DB_PATH)
    conn.execute("DELETE FROM expenses WHERE id = ?", (expense_id,))
    conn.commit()
    conn.close()

# ── Pydantic model for OCR output ─────────────────────────────────────────────
class BillData(BaseModel):
    expense_date: str = Field(description="Date on the bill in YYYY-MM-DD format. Use today if not found.")
    vendor_name: str = Field(description="Vendor or shop name from the bill")
    particulars: str = Field(description="Brief description of what was purchased")
    total_amount: float = Field(description="Total amount paid in rupees")
    payment_mode: str = Field(description="Cash, UPI, Card, or Unknown")
    category: str = Field(description="One of: Medicine, Hardware, Transport, Food/Tea, Stationery, Electricity, Rent, Salary, Misc")

# ── AI Bill Reading ────────────────────────────────────────────────────────────
def read_bill_with_ai(image_bytes: bytes) -> Optional[BillData]:
    if not GEMINI_API_KEY:
        st.error("Gemini API key not set. Please add GEMINI_API_KEY in Streamlit secrets.")
        return None
    try:
        model = genai.GenerativeModel("gemini-1.5-flash")
        img = Image.open(BytesIO(image_bytes))
        prompt = """You are an expert at reading Indian business bills, receipts, and transport challans.
Extract the following from this bill image and return ONLY valid JSON:
{
  "expense_date": "YYYY-MM-DD",
  "vendor_name": "shop or vendor name",
  "particulars": "brief description of purchase",
  "total_amount": 0.00,
  "payment_mode": "Cash or UPI or Card or Unknown",
  "category": "one of: Medicine, Hardware, Transport, Food/Tea, Stationery, Electricity, Rent, Salary, Misc"
}
If date is not visible, use today's date. If amount is not clear, use 0. Return ONLY the JSON, no explanation."""
        response = model.generate_content([prompt, img])
        text = response.text.strip()
        # Clean up markdown code blocks if present
        if text.startswith("```"):
            text = text.split("```")[1]
            if text.startswith("json"):
                text = text[4:]
        data = json.loads(text.strip())
        return BillData(**data)
    except Exception as e:
        st.error(f"AI reading failed: {e}")
        return None

# ── Excel Export ───────────────────────────────────────────────────────────────
def export_to_excel(df: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Prepare export dataframe
        export_df = df.copy()
        export_df = export_df.rename(columns={
            "expense_date": "Date",
            "particulars": "Particulars",
            "vendor": "Vendor Name",
            "category": "Category",
            "amount": "Amount (Rs.)",
            "payment_mode": "Payment Mode",
            "entered_by": "Entered By",
            "notes": "Notes",
        })
        cols = ["Date", "Vendor Name", "Particulars", "Category", "Amount (Rs.)", "Payment Mode", "Entered By", "Notes"]
        export_df = export_df[[c for c in cols if c in export_df.columns]]

        export_df.to_excel(writer, index=False, sheet_name="Expenses")

        # Style the sheet
        ws = writer.sheets["Expenses"]
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="17533A", end_color="17533A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Add company name header row above data
        ws.insert_rows(1)
        ws.insert_rows(1)
        ws["A1"] = "KALE PHARMA PVT LTD — Expense Statement"
        ws["A1"].font = Font(bold=True, size=14, color="17533A")
        ws["A2"] = f"Generated on: {datetime.now().strftime('%d/%m/%Y %I:%M %p')}"
        ws["A2"].font = Font(italic=True, size=10, color="666666")

        # Auto-fit columns
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    return output.getvalue()

# ── Categories ─────────────────────────────────────────────────────────────────
CATEGORIES = ["Medicine", "Hardware", "Transport", "Food/Tea", "Stationery", "Electricity", "Rent", "Salary", "Misc"]
PAYMENT_MODES = ["Cash", "UPI", "Card", "Cheque", "Other"]

# ── Init ───────────────────────────────────────────────────────────────────────
init_db()

# ── Sidebar Navigation ─────────────────────────────────────────────────────────
st.sidebar.markdown("## 💊 Swara")
st.sidebar.markdown("**Kale Pharma Pvt Ltd**")
st.sidebar.markdown("---")
page = st.sidebar.radio(
    "Navigate",
    ["➕ Add Expense", "📋 Statement", "📊 Summary"],
    label_visibility="collapsed"
)
st.sidebar.markdown("---")
st.sidebar.markdown("*Expense Tracker v1.0*")

# ══════════════════════════════════════════════════════════════════════════════
# PAGE: ADD EXPENSE
# ══════════════════════════════════════════════════════════════════════════════
if page == "➕ Add Expense":
    st.markdown('<p class="company-header">Kale Pharma Pvt Ltd</p>', unsafe_allow_html=True)
    st.markdown('<h1 class="page-title">Add Expense</h1>', unsafe_allow_html=True)
    st.caption("Upload a bill photo — Swara reads the details for you, or enter manually.")

    tab1, tab2 = st.tabs(["📷 Upload Bill Photo", "✏️ Enter Manually"])

    with tab1:
        uploaded_file = st.file_uploader(
            "Upload bill photo (receipt, transport bilty, handwritten slip)",
            type=["jpg", "jpeg", "png", "webp"],
            label_visibility="collapsed"
        )
        if uploaded_file:
            col1, col2 = st.columns([1, 2])
            with col1:
                st.image(uploaded_file, caption="Uploaded Bill", use_column_width=True)
            with col2:
                if st.button("🤖 Read Bill with AI", use_container_width=True):
                    with st.spinner("Reading bill..."):
                        bill_data = read_bill_with_ai(uploaded_file.read())
                    if bill_data:
                        st.session_state["ai_bill"] = bill_data
                        st.success("Bill read successfully! Review and save below.")

            if "ai_bill" in st.session_state:
                bill = st.session_state["ai_bill"]
                st.markdown("#### Review & Save")
                with st.form("ai_expense_form"):
                    c1, c2 = st.columns(2)
                    exp_date = c1.date_input("Date", value=datetime.strptime(bill.expense_date, "%Y-%m-%d").date() if bill.expense_date else date.today())
                    vendor = c2.text_input("Vendor Name", value=bill.vendor_name)
                    particulars = st.text_input("Particulars", value=bill.particulars)
                    c3, c4, c5 = st.columns(3)
                    amount = c3.number_input("Amount (Rs.)", value=float(bill.total_amount), min_value=0.0, step=0.5)
                    category = c4.selectbox("Category", CATEGORIES, index=CATEGORIES.index(bill.category) if bill.category in CATEGORIES else 0)
                    payment_mode = c5.selectbox("Payment Mode", PAYMENT_MODES, index=PAYMENT_MODES.index(bill.payment_mode) if bill.payment_mode in PAYMENT_MODES else 0)
                    entered_by = st.text_input("Entered By", value="Owner")
                    notes = st.text_area("Notes (optional)", height=60)
                    if st.form_submit_button("💾 Save Expense", use_container_width=True):
                        add_expense(str(exp_date), particulars, vendor, category, amount, payment_mode, entered_by, notes)
                        del st.session_state["ai_bill"]
                        st.success(f"✅ Expense of ₹{amount:.0f} saved!")
                        st.rerun()

    with tab2:
        with st.form("manual_expense_form"):
            c1, c2 = st.columns(2)
            exp_date = c1.date_input("Date", value=date.today())
            vendor = c2.text_input("Vendor Name")
            particulars = st.text_input("Particulars *", placeholder="e.g. Purchased medicines from supplier")
            c3, c4, c5 = st.columns(3)
            amount = c3.number_input("Amount (Rs.) *", min_value=0.0, step=0.5)
            category = c4.selectbox("Category", CATEGORIES)
            payment_mode = c5.selectbox("Payment Mode", PAYMENT_MODES)
            entered_by = st.text_input("Entered By", value="Owner")
            notes = st.text_area("Notes (optional)", height=60)
            if st.form_submit_button("💾 Save Expense", use_container_width=True):
                if not particulars:
                    st.error("Please enter Particulars.")
                elif amount <= 0:
                    st.error("Please enter a valid amount.")
                else:
                    add_expense(str(exp_date), particulars, vendor, category, amount, payment_mode, entered_by, notes)
                    st.success(f"✅ Expense of ₹{amount:.0f} saved!")
                    st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# PAGE: STATEMENT
# ══════════════════════════════════════════════════════════════════════════════
elif page == "📋 Statement":
    st.markdown('<p class="company-header">Kale Pharma Pvt Ltd</p>', unsafe_allow_html=True)
    st.markdown('<h1 class="page-title">Expense Statement</h1>', unsafe_allow_html=True)

    df = get_all_expenses()

    # Filters
    col1, col2, col3 = st.columns(3)
    today = date.today()
    from_date = col1.date_input("From", value=date(today.year, today.month, 1))
    to_date = col2.date_input("To", value=today)
    cat_filter = col3.selectbox("Category", ["All"] + CATEGORIES)

    if not df.empty:
        df["expense_date"] = pd.to_datetime(df["expense_date"]).dt.date
        mask = (df["expense_date"] >= from_date) & (df["expense_date"] <= to_date)
        if cat_filter != "All":
            mask &= df["category"] == cat_filter
        filtered = df[mask].copy()
    else:
        filtered = df.copy()

    # Summary metrics
    total = filtered["amount"].sum() if not filtered.empty else 0
    count = len(filtered)
    m1, m2, m3 = st.columns(3)
    m1.metric("Total Spend", f"₹ {total:,.0f}")
    m2.metric("Entries", count)
    m3.metric("Avg per Entry", f"₹ {total/count:,.0f}" if count > 0 else "₹ 0")

    st.markdown("---")

    # Export buttons
    if not filtered.empty:
        ec1, ec2 = st.columns([1, 1])
        with ec1:
            excel_bytes = export_to_excel(filtered)
            st.download_button(
                label="📥 Download Excel",
                data=excel_bytes,
                file_name=f"kale-pharma-expenses-{from_date}-to-{to_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    # Table
    if filtered.empty:
        st.info("No expenses found for the selected period.")
    else:
        display_df = filtered[["expense_date", "vendor", "particulars", "category", "amount", "payment_mode", "entered_by"]].copy()
        display_df.columns = ["Date", "Vendor", "Particulars", "Category", "Amount (₹)", "Payment", "By"]
        display_df["Amount (₹)"] = display_df["Amount (₹)"].apply(lambda x: f"₹ {x:,.0f}")
        st.dataframe(display_df, use_container_width=True, hide_index=True)

        # Delete option
        st.markdown("#### Delete an Entry")
        del_id = st.number_input("Enter ID to delete", min_value=1, step=1, value=1)
        if st.button("🗑️ Delete Entry", type="secondary"):
            delete_expense(int(del_id))
            st.success(f"Entry #{del_id} deleted.")
            st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# PAGE: SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
elif page == "📊 Summary":
    st.markdown('<p class="company-header">Kale Pharma Pvt Ltd</p>', unsafe_allow_html=True)
    st.markdown('<h1 class="page-title">Summary</h1>', unsafe_allow_html=True)
    st.caption("Spending totals and category breakdown.")

    df = get_all_expenses()
    today = date.today()

    period = st.radio("Period", ["Today", "This Month", "This Year", "All Time"], horizontal=True)
    if period == "Today":
        mask = pd.to_datetime(df["expense_date"]).dt.date == today if not df.empty else pd.Series([], dtype=bool)
    elif period == "This Month":
        mask = (pd.to_datetime(df["expense_date"]).dt.year == today.year) & \
               (pd.to_datetime(df["expense_date"]).dt.month == today.month) if not df.empty else pd.Series([], dtype=bool)
    elif period == "This Year":
        mask = pd.to_datetime(df["expense_date"]).dt.year == today.year if not df.empty else pd.Series([], dtype=bool)
    else:
        mask = pd.Series([True] * len(df))

    filtered = df[mask].copy() if not df.empty and len(mask) > 0 else pd.DataFrame()

    total = filtered["amount"].sum() if not filtered.empty else 0
    count = len(filtered)

    m1, m2, m3 = st.columns(3)
    m1.metric("Total Spend", f"₹ {total:,.0f}")
    m2.metric("Entries", count)
    m3.metric("Avg per Entry", f"₹ {total/count:,.0f}" if count > 0 else "₹ 0")

    if not filtered.empty:
        st.markdown("---")
        col1, col2 = st.columns(2)

        with col1:
            st.markdown("#### By Category")
            cat_summary = filtered.groupby("category")["amount"].sum().reset_index()
            cat_summary.columns = ["Category", "Amount (₹)"]
            cat_summary = cat_summary.sort_values("Amount (₹)", ascending=False)
            cat_summary["Amount (₹)"] = cat_summary["Amount (₹)"].apply(lambda x: f"₹ {x:,.0f}")
            st.dataframe(cat_summary, use_container_width=True, hide_index=True)

        with col2:
            st.markdown("#### By Payment Mode")
            pay_summary = filtered.groupby("payment_mode")["amount"].sum().reset_index()
            pay_summary.columns = ["Payment Mode", "Amount (₹)"]
            pay_summary = pay_summary.sort_values("Amount (₹)", ascending=False)
            pay_summary["Amount (₹)"] = pay_summary["Amount (₹)"].apply(lambda x: f"₹ {x:,.0f}")
            st.dataframe(pay_summary, use_container_width=True, hide_index=True)

        st.markdown("#### Daily Spend")
        daily = filtered.copy()
        daily["expense_date"] = pd.to_datetime(daily["expense_date"])
        daily_sum = daily.groupby("expense_date")["amount"].sum().reset_index()
        daily_sum.columns = ["Date", "Amount (₹)"]
        st.bar_chart(daily_sum.set_index("Date"))
    else:
        st.info("No expenses found for the selected period.")
